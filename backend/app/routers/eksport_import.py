import io
import csv
import logging
from datetime import date
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import Wydatek, Splata
from ..schemas import MessageResponse
from ..auth import verify_password

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api", tags=["eksport-import"])

OSOBY_VALID = {"mateusz", "jan", "wojciech"}


# ─── Eksport CSV ─────────────────────────────────────────────────────────────

@router.get("/eksport/csv")
def eksport_csv(db: Session = Depends(get_db), _: str = Depends(verify_password)):
    wydatki = db.query(Wydatek).order_by(Wydatek.data).all()
    splaty = db.query(Splata).order_by(Splata.data).all()

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")

    writer.writerow(["=== WYDATKI ==="])
    writer.writerow(["ID", "Data", "Opis", "Fundator", "Udział Mateusz", "Udział Jan", "Udział Wojciech", "Razem"])
    for w in wydatki:
        razem = (w.udzial_mateusz or 0) + (w.udzial_jan or 0) + (w.udzial_wojciech or 0)
        writer.writerow([
            w.id,
            w.data.strftime("%d.%m.%Y") if w.data else "",
            w.opis,
            w.fundator,
            str(w.udzial_mateusz or 0).replace(".", ","),
            str(w.udzial_jan or 0).replace(".", ","),
            str(w.udzial_wojciech or 0).replace(".", ","),
            str(razem).replace(".", ","),
        ])

    writer.writerow([])
    writer.writerow(["=== SPŁATY ==="])
    writer.writerow(["ID", "Data", "Kto", "Kwota"])
    for s in splaty:
        writer.writerow([
            s.id,
            s.data.strftime("%d.%m.%Y") if s.data else "",
            s.kto,
            str(s.kwota).replace(".", ","),
        ])

    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": "attachment; filename=rozliczenia.csv"},
    )


# ─── Eksport XLSX ─────────────────────────────────────────────────────────────

@router.get("/eksport/xlsx")
def eksport_xlsx(db: Session = Depends(get_db), _: str = Depends(verify_password)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl nie zainstalowany")

    wydatki = db.query(Wydatek).order_by(Wydatek.data).all()
    splaty = db.query(Splata).order_by(Splata.data).all()

    wb = openpyxl.Workbook()

    # Arkusz Wydatki
    ws_w = wb.active
    ws_w.title = "Wydatki"
    headers_w = ["ID", "Data", "Opis", "Fundator", "Udział Mateusz", "Udział Jan", "Udział Wojciech", "Razem"]
    ws_w.append(headers_w)
    header_font = Font(bold=True)
    for cell in ws_w[1]:
        cell.font = header_font

    for w in wydatki:
        razem = float((w.udzial_mateusz or 0) + (w.udzial_jan or 0) + (w.udzial_wojciech or 0))
        ws_w.append([
            w.id,
            w.data,
            w.opis,
            w.fundator,
            float(w.udzial_mateusz or 0),
            float(w.udzial_jan or 0),
            float(w.udzial_wojciech or 0),
            razem,
        ])

    # Arkusz Splaty
    ws_s = wb.create_sheet("Splaty")
    headers_s = ["ID", "Data", "Kto", "Kwota"]
    ws_s.append(headers_s)
    for cell in ws_s[1]:
        cell.font = header_font

    for s in splaty:
        ws_s.append([s.id, s.data, s.kto, float(s.kwota)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rozliczenia.xlsx"},
    )


# ─── Import XLSX ──────────────────────────────────────────────────────────────

@router.post("/import/xlsx", response_model=MessageResponse)
async def import_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: str = Depends(verify_password),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl nie zainstalowany")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Obsługiwane formaty: .xlsx, .xls")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Błąd odczytu pliku: {exc}")

    stats = {"wydatki_dodane": 0, "splaty_dodane": 0, "bledy": []}

    # Import arkusza Wydatki
    if "Wydatki" in wb.sheetnames:
        ws = wb["Wydatki"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row_num, row in enumerate(rows, 2):
            if not any(row):
                continue
            try:
                _, data_raw, opis, fundator, udz_m, udz_j, udz_w = (list(row) + [None] * 7)[:7]
                if not data_raw or not opis or not fundator:
                    continue

                if isinstance(data_raw, str):
                    from datetime import datetime as dt
                    data_val = dt.strptime(data_raw.strip(), "%d.%m.%Y").date()
                elif hasattr(data_raw, "date"):
                    data_val = data_raw.date() if hasattr(data_raw, "date") else data_raw
                else:
                    data_val = data_raw

                fundator_norm = str(fundator).strip().lower()
                if fundator_norm not in OSOBY_VALID:
                    stats["bledy"].append(f"Wiersz {row_num}: nieznany fundator '{fundator}'")
                    continue

                w = Wydatek(
                    data=data_val,
                    opis=str(opis),
                    fundator=fundator_norm.capitalize(),
                    udzial_mateusz=float(udz_m or 0),
                    udzial_jan=float(udz_j or 0),
                    udzial_wojciech=float(udz_w or 0),
                )
                db.add(w)
                stats["wydatki_dodane"] += 1
            except Exception as exc:
                stats["bledy"].append(f"Wiersz {row_num} (Wydatki): {exc}")

    # Import arkusza Splaty
    if "Splaty" in wb.sheetnames:
        ws = wb["Splaty"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row_num, row in enumerate(rows, 2):
            if not any(row):
                continue
            try:
                _, data_raw, kto, kwota = (list(row) + [None] * 4)[:4]
                if not data_raw or not kto or not kwota:
                    continue

                if isinstance(data_raw, str):
                    from datetime import datetime as dt
                    data_val = dt.strptime(data_raw.strip(), "%d.%m.%Y").date()
                elif hasattr(data_raw, "date"):
                    data_val = data_raw.date() if hasattr(data_raw, "date") else data_raw
                else:
                    data_val = data_raw

                kto_norm = str(kto).strip().lower()
                if kto_norm not in OSOBY_VALID:
                    stats["bledy"].append(f"Wiersz {row_num}: nieznana osoba '{kto}'")
                    continue

                s = Splata(
                    data=data_val,
                    kto=kto_norm.capitalize(),
                    kwota=float(kwota),
                )
                db.add(s)
                stats["splaty_dodane"] += 1
            except Exception as exc:
                stats["bledy"].append(f"Wiersz {row_num} (Splaty): {exc}")

    db.commit()
    return MessageResponse(
        message=f"Import zakończony: {stats['wydatki_dodane']} wydatków, {stats['splaty_dodane']} spłat",
        details=stats,
    )
